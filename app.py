import os
import io
import json
from collections import defaultdict
from math import floor
from datetime import datetime, timedelta, date
from functools import wraps

from flask import (Flask, render_template, request, redirect, url_for,
                   session, jsonify, send_file, flash)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# APP CONFIG
# ============================================================
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'copiadora_peru_2024_xK9mL2')
# Render usa 'postgres://' pero SQLAlchemy necesita 'postgresql://'
_db_url = os.environ.get('DATABASE_URL', 'sqlite:///copiadora.db')
if _db_url.startswith('postgres://'):
    _db_url = _db_url.replace('postgres://', 'postgresql+psycopg://', 1)
elif _db_url.startswith('postgresql://') and '+psycopg' not in _db_url:
    _db_url = _db_url.replace('postgresql://', 'postgresql+psycopg://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)

db = SQLAlchemy(app)

# ============================================================
# MODELOS
# ============================================================

class Usuario(db.Model):
    __tablename__ = 'usuarios'
    id         = db.Column(db.Integer, primary_key=True)
    nombre     = db.Column(db.String(100), nullable=False)
    pin        = db.Column(db.String(256), nullable=False)
    rol        = db.Column(db.String(20), default='apoyo')   # admin / apoyo
    activo     = db.Column(db.Boolean, default=True)


class Servicio(db.Model):
    __tablename__ = 'servicios'
    id                  = db.Column(db.Integer, primary_key=True)
    nombre              = db.Column(db.String(100), nullable=False)
    precio              = db.Column(db.Float, nullable=False)
    costo_real          = db.Column(db.Float, nullable=False)
    activo              = db.Column(db.Boolean, default=True)
    descuento_volumen   = db.Column(db.Boolean, default=False)  # 8% en copias/impresiones


class ServicioNivel(db.Model):
    __tablename__ = 'servicio_niveles'
    id          = db.Column(db.Integer, primary_key=True)
    servicio_id = db.Column(db.Integer, db.ForeignKey('servicios.id'), nullable=False)
    nombre      = db.Column(db.String(100), nullable=False)
    precio      = db.Column(db.Float, nullable=False)
    orden       = db.Column(db.Integer, default=0)

    servicio = db.relationship('Servicio', backref='niveles')


class Cliente(db.Model):
    __tablename__ = 'clientes'
    id              = db.Column(db.Integer, primary_key=True)
    nombre          = db.Column(db.String(100), nullable=False)
    celular         = db.Column(db.String(20), nullable=True)
    puntos          = db.Column(db.Integer, default=0)
    activo          = db.Column(db.Boolean, default=True)
    fecha_registro  = db.Column(db.Date, default=date.today)


class Venta(db.Model):
    __tablename__ = 'ventas'
    id             = db.Column(db.Integer, primary_key=True)
    fecha          = db.Column(db.Date, nullable=False)
    hora           = db.Column(db.Time, nullable=False)
    servicio_id    = db.Column(db.Integer, db.ForeignKey('servicios.id'), nullable=False)
    cantidad       = db.Column(db.Integer, nullable=False)
    precio_unitario= db.Column(db.Float, nullable=False)
    total          = db.Column(db.Float, nullable=False)
    usuario_id     = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    cliente_id     = db.Column(db.Integer, db.ForeignKey('clientes.id'), nullable=True)
    es_canje       = db.Column(db.Boolean, default=False)
    descripcion    = db.Column(db.String(200), nullable=True)
    nivel_nombre   = db.Column(db.String(100), nullable=True)
    nota           = db.Column(db.String(200), nullable=True)
    descuento      = db.Column(db.Float, default=0)

    servicio = db.relationship('Servicio', backref='ventas')
    usuario  = db.relationship('Usuario',  backref='ventas')
    cliente  = db.relationship('Cliente',  backref='ventas')


class Inventario(db.Model):
    __tablename__ = 'inventario'
    id       = db.Column(db.Integer, primary_key=True)
    fecha    = db.Column(db.Date, nullable=False)
    tipo     = db.Column(db.String(50), nullable=False)   # papel / tinta_negra / tinta_color
    cantidad = db.Column(db.Integer, nullable=False)
    nota     = db.Column(db.String(200), nullable=True)


class Configuracion(db.Model):
    __tablename__ = 'configuracion'
    id    = db.Column(db.Integer, primary_key=True)
    clave = db.Column(db.String(50), unique=True, nullable=False)
    valor = db.Column(db.String(200), nullable=False)


class PuntosHistorial(db.Model):
    __tablename__ = 'puntos_historial'
    id          = db.Column(db.Integer, primary_key=True)
    cliente_id  = db.Column(db.Integer, db.ForeignKey('clientes.id'), nullable=False)
    fecha       = db.Column(db.DateTime, default=datetime.now)
    puntos      = db.Column(db.Integer, nullable=False)
    descripcion = db.Column(db.String(200))

    cliente = db.relationship('Cliente', backref='historial_puntos')


# ============================================================
# HELPERS
# ============================================================

def get_config(clave, default=None):
    c = Configuracion.query.filter_by(clave=clave).first()
    return c.valor if c else default


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect(url_for('login'))
        if session.get('rol') != 'admin':
            return redirect(url_for('ventas'))
        return f(*args, **kwargs)
    return decorated


# ============================================================
# INICIALIZAR BASE DE DATOS
# ============================================================

def init_db():
    db.create_all()

    # Migraciones — SIEMPRE primero, antes de cualquier query
    migraciones = [
        'ALTER TABLE ventas ADD COLUMN IF NOT EXISTS descripcion VARCHAR(200)',
        'ALTER TABLE ventas ADD COLUMN IF NOT EXISTS nivel_nombre VARCHAR(100)',
        'ALTER TABLE ventas ADD COLUMN IF NOT EXISTS nota VARCHAR(200)',
        'ALTER TABLE ventas ADD COLUMN IF NOT EXISTS descuento FLOAT DEFAULT 0',
        'ALTER TABLE servicios ADD COLUMN IF NOT EXISTS descuento_volumen BOOLEAN DEFAULT FALSE',
    ]
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            for sql in migraciones:
                try:
                    conn.execute(text(sql))
                except Exception:
                    pass
            conn.commit()
    except Exception:
        pass

    # Usuarios
    if not Usuario.query.first():
        db.session.add_all([
            Usuario(nombre='Admin', pin=generate_password_hash('1234'), rol='admin'),
            Usuario(nombre='Apoyo', pin=generate_password_hash('0000'), rol='apoyo'),
        ])

    # Servicios con niveles de precio
    if not Servicio.query.first():
        servicios_seed = [
            ('Copia B&N',          0.10, 0.023, True,  [('Texto', 0.10), ('Con imagen', 0.20)]),
            ('Copia color',        0.30, 0.034, True,  [('Simple', 0.30), ('Densa', 0.50)]),
            ('Impresión B&N',      0.30, 0.026, True,  [('Texto', 0.30), ('Con imágenes', 0.50)]),
            ('Impresión color',    0.50, 0.034, True,  [('Normal', 0.50), ('Alta cobertura', 1.00)]),
            ('Escaneo',            1.00, 0.010, False, [('Página suelta', 1.00), ('Doc. completo ≤5 págs', 2.00)]),
            ('Foto carné (x6)',    4.00, 0.034, False, []),
            ('Llenado formulario', 2.50, 0.026, False, []),
        ]
        for nombre, precio, costo, desc_vol, niveles in servicios_seed:
            s = Servicio(nombre=nombre, precio=precio, costo_real=costo,
                         descuento_volumen=desc_vol)
            db.session.add(s)
            db.session.flush()
            for i, (nnom, nprec) in enumerate(niveles):
                db.session.add(ServicioNivel(servicio_id=s.id, nombre=nnom,
                                             precio=nprec, orden=i))

    # Servicio especial para ventas personalizadas (oculto en botones)
    if not Servicio.query.filter_by(nombre='__personalizado__').first():
        db.session.add(Servicio(nombre='__personalizado__', precio=0,
                                costo_real=0, activo=False))

    # Inventario inicial
    if not Inventario.query.first():
        db.session.add(Inventario(fecha=date.today(), tipo='papel', cantidad=500,
                                  nota='Stock inicial'))

    # Configuración
    defaults = {
        'hojas_iniciales':    '500',
        'alerta_minimo':      '100',
        'nombre_negocio':     'Mi Copiadora',
        'costos_fijos_mes':   '160',
    }
    for clave, valor in defaults.items():
        if not Configuracion.query.filter_by(clave=clave).first():
            db.session.add(Configuracion(clave=clave, valor=valor))

    db.session.commit()


# ============================================================
# CONTEXT PROCESSOR
# ============================================================

@app.context_processor
def inject_globals():
    return {'get_config': get_config}


# ============================================================
# RUTAS GENERALES
# ============================================================

@app.route('/')
def index():
    if 'usuario_id' in session:
        return redirect(url_for('panel') if session.get('rol') == 'admin' else url_for('ventas'))
    return redirect(url_for('login'))


@app.route('/ping')
def ping():
    return jsonify({'status': 'ok'})


# ============================================================
# AUTENTICACIÓN
# ============================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        nombre_ingresado = request.form.get('nombre', '').strip().lower()
        pin              = request.form.get('pin', '')
        if not nombre_ingresado:
            error = 'Ingresa tu nombre de usuario.'
        else:
            usuario = Usuario.query.filter(
                db.func.lower(Usuario.nombre) == nombre_ingresado,
                Usuario.activo == True
            ).first()
            if usuario and check_password_hash(usuario.pin, pin):
                session.permanent     = True
                session['usuario_id'] = usuario.id
                session['nombre']     = usuario.nombre
                session['rol']        = usuario.rol
                return redirect(url_for('panel') if usuario.rol == 'admin' else url_for('ventas'))
            else:
                error = 'Usuario o PIN incorrecto. Intenta de nuevo.'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ============================================================
# VENTAS
# ============================================================

@app.route('/ventas')
@login_required
def ventas():
    from sqlalchemy import func
    servicios = Servicio.query.filter_by(activo=True).all()
    hoy = date.today()

    ultimas = (Venta.query
               .filter_by(fecha=hoy, usuario_id=session['usuario_id'])
               .order_by(Venta.id.desc()).limit(5).all())

    total_turno = (db.session.query(func.sum(Venta.total))
                   .filter_by(fecha=hoy, usuario_id=session['usuario_id'])
                   .scalar() or 0)

    clientes = Cliente.query.filter_by(activo=True).order_by(Cliente.nombre).all()

    return render_template('ventas.html',
                           servicios=servicios,
                           ultimas=ultimas,
                           total_turno=total_turno,
                           clientes=clientes)


@app.route('/ventas/registrar', methods=['POST'])
@login_required
def registrar_venta():
    servicio_id = request.form.get('servicio_id', type=int)
    cantidad    = request.form.get('cantidad',    type=int)
    cliente_id  = request.form.get('cliente_id', type=int) or None

    if not servicio_id or not cantidad or cantidad <= 0:
        return jsonify({'error': 'Datos inválidos'}), 400

    srv = Servicio.query.get(servicio_id)
    if not srv or not srv.activo:
        return jsonify({'error': 'Servicio no disponible'}), 400

    nivel_id     = request.form.get('nivel_id', type=int)
    precio_manual= request.form.get('precio_manual', type=float)
    nota         = request.form.get('nota', '').strip() or None
    nivel_nombre = None

    # Determinar precio: nivel > manual > servicio base
    if nivel_id:
        niv = ServicioNivel.query.get(nivel_id)
        if niv and niv.servicio_id == servicio_id:
            precio_final = niv.precio
            nivel_nombre = niv.nombre
        else:
            precio_final = srv.precio
    elif precio_manual and precio_manual >= 0:
        precio_final = precio_manual
    elif srv.precio == 0 and not precio_manual:
        return jsonify({'error': 'Este servicio requiere un precio manual'}), 400
    else:
        precio_final = srv.precio

    # Descuento por volumen: 8% en copias/impresiones con 10+ unidades
    UMBRAL_DESCUENTO = 10
    descuento = 0.0
    if srv.descuento_volumen and cantidad >= UMBRAL_DESCUENTO:
        descuento = 8.0

    subtotal  = precio_final * cantidad
    total     = round(subtotal * (1 - descuento / 100), 2)

    now = datetime.now()
    v = Venta(fecha=now.date(), hora=now.time(),
              servicio_id=servicio_id, cantidad=cantidad,
              precio_unitario=precio_final, total=total,
              usuario_id=session['usuario_id'], cliente_id=cliente_id,
              nivel_nombre=nivel_nombre, nota=nota, descuento=descuento,
              es_canje=False)
    db.session.add(v)

    puntos_ganados = 0
    puntos_actuales = 0
    if cliente_id:
        cli = Cliente.query.get(cliente_id)
        if cli:
            puntos_ganados = floor(total / 10)
            if puntos_ganados > 0:
                cli.puntos += puntos_ganados
                db.session.add(PuntosHistorial(
                    cliente_id=cliente_id, puntos=puntos_ganados,
                    descripcion=f'Compra: {srv.nombre} x{cantidad} = S/{total:.2f}'))
            puntos_actuales = cli.puntos + puntos_ganados

    db.session.commit()
    return jsonify({
        'success':         True,
        'total':           total,
        'puntos_ganados':  puntos_ganados,
        'puntos_actuales': puntos_actuales,
        'servicio':        srv.nombre + (f' ({nivel_nombre})' if nivel_nombre else ''),
        'cantidad':        cantidad,
        'precio_variable': srv.precio == 0,
        'descuento':       descuento,
    })


@app.route('/ventas/personalizada', methods=['POST'])
@login_required
def venta_personalizada():
    descripcion = request.form.get('descripcion', '').strip()
    precio      = request.form.get('precio',      type=float)
    cantidad    = request.form.get('cantidad',    type=int)
    cliente_id  = request.form.get('cliente_id', type=int) or None
    usa_hojas   = request.form.get('usa_hojas', 'true') == 'true'

    if not descripcion:
        return jsonify({'error': 'Describe el servicio'}), 400
    if precio is None or precio < 0:
        return jsonify({'error': 'Precio inválido'}), 400
    if not cantidad or cantidad <= 0:
        return jsonify({'error': 'Cantidad inválida'}), 400

    # Elegir servicio según si usa hojas o no
    srv_nombre = '__personalizado__' if usa_hojas else '__personalizado_sinhojas__'
    srv = Servicio.query.filter_by(nombre=srv_nombre).first()
    if not srv:
        srv = Servicio(nombre=srv_nombre, precio=0, costo_real=0, activo=False)
        db.session.add(srv)
        db.session.flush()

    now   = datetime.now()
    total = round(precio * cantidad, 2)

    v = Venta(fecha=now.date(), hora=now.time(),
              servicio_id=srv.id, cantidad=cantidad,
              precio_unitario=precio, total=total,
              usuario_id=session['usuario_id'], cliente_id=cliente_id,
              descripcion=descripcion, es_canje=False)
    db.session.add(v)

    puntos_ganados  = 0
    puntos_actuales = 0
    if cliente_id:
        cli = Cliente.query.get(cliente_id)
        if cli:
            puntos_ganados = floor(total / 10)
            if puntos_ganados > 0:
                cli.puntos += puntos_ganados
                db.session.add(PuntosHistorial(
                    cliente_id=cliente_id, puntos=puntos_ganados,
                    descripcion=f'Compra: {descripcion} x{cantidad} = S/{total:.2f}'))
            puntos_actuales = cli.puntos + puntos_ganados

    db.session.commit()
    return jsonify({
        'success':         True,
        'total':           total,
        'puntos_ganados':  puntos_ganados,
        'puntos_actuales': puntos_actuales,
        'servicio':        descripcion,
        'cantidad':        cantidad,
    })


@app.route('/ventas/canjear', methods=['POST'])
@login_required
def canjear_puntos():
    cliente_id = request.form.get('cliente_id', type=int)
    if not cliente_id:
        return jsonify({'error': 'Cliente no especificado'}), 400

    cli = Cliente.query.get(cliente_id)
    if not cli or cli.puntos < 10:
        return jsonify({'error': 'Puntos insuficientes (mínimo 10)'}), 400

    srv = Servicio.query.filter(Servicio.nombre.like('%Copia B&N texto%')).first() \
          or Servicio.query.first()

    now = datetime.now()
    v = Venta(fecha=now.date(), hora=now.time(),
              servicio_id=srv.id, cantidad=1,
              precio_unitario=0, total=0,
              usuario_id=session['usuario_id'], cliente_id=cliente_id,
              es_canje=True)
    db.session.add(v)

    cli.puntos -= 10
    db.session.add(PuntosHistorial(
        cliente_id=cliente_id, puntos=-10,
        descripcion='Canje: 1 copia gratis'))
    db.session.commit()

    return jsonify({'success': True, 'puntos_restantes': cli.puntos})


# ============================================================
# PANEL
# ============================================================

@app.route('/panel')
@admin_required
def panel():
    from sqlalchemy import func
    hoy         = date.today()
    lunes       = hoy - timedelta(days=hoy.weekday())

    ventas_hoy  = Venta.query.filter_by(fecha=hoy).all()
    ing_hoy     = sum(v.total for v in ventas_hoy)
    tx_hoy      = len(ventas_hoy)

    top = (db.session.query(Servicio.nombre, func.sum(Venta.cantidad).label('tot'))
           .join(Venta).filter(Venta.fecha == hoy)
           .group_by(Servicio.id).order_by(func.sum(Venta.cantidad).desc()).first())
    srv_top = top[0] if top else '—'

    grafico = (db.session.query(Servicio.nombre,
                                func.sum(Venta.cantidad).label('cant'),
                                func.sum(Venta.total).label('ing'))
               .join(Venta).filter(Venta.fecha == hoy)
               .group_by(Servicio.id).all())

    grafico_json = json.dumps([
        {'nombre': r[0], 'cantidad': int(r[1] or 0), 'ingresos': float(r[2] or 0)}
        for r in grafico
    ])

    ventas_sem = Venta.query.filter(Venta.fecha >= lunes, Venta.fecha <= hoy).all()
    ing_sem    = sum(v.total for v in ventas_sem)
    tx_sem     = len(ventas_sem)

    nombre_negocio = get_config('nombre_negocio', 'Mi Copiadora')

    return render_template('panel.html',
                           ing_hoy=ing_hoy, tx_hoy=tx_hoy, srv_top=srv_top,
                           grafico_json=grafico_json,
                           ing_sem=ing_sem, tx_sem=tx_sem,
                           nombre_negocio=nombre_negocio, hoy=hoy)


# ============================================================
# HISTORIAL
# ============================================================

@app.route('/historial')
@admin_required
def historial():
    desde      = request.args.get('desde', '')
    hasta      = request.args.get('hasta', '')
    srv_id     = request.args.get('servicio_id', '')
    pagina     = request.args.get('pagina', 1, type=int)
    por_pagina = 20

    q = Venta.query
    if desde:
        q = q.filter(Venta.fecha >= datetime.strptime(desde, '%Y-%m-%d').date())
    if hasta:
        q = q.filter(Venta.fecha <= datetime.strptime(hasta, '%Y-%m-%d').date())
    if srv_id and srv_id.isdigit():
        q = q.filter(Venta.servicio_id == int(srv_id))
    q = q.order_by(Venta.fecha.desc(), Venta.hora.desc())

    total_count  = q.count()
    ventas       = q.offset((pagina - 1) * por_pagina).limit(por_pagina).all()
    total_paginas= (total_count + por_pagina - 1) // por_pagina
    total_monto  = sum(v.total for v in q.all())

    servicios = Servicio.query.all()

    return render_template('historial.html',
                           ventas=ventas, servicios=servicios,
                           desde=desde, hasta=hasta, srv_id=srv_id,
                           pagina=pagina, total_paginas=total_paginas,
                           total_monto=total_monto, total_count=total_count)


# ============================================================
# INVENTARIO
# ============================================================

@app.route('/inventario')
@admin_required
def inventario():
    from sqlalchemy import func
    total_papel  = (db.session.query(func.sum(Inventario.cantidad))
                    .filter_by(tipo='papel').scalar() or 0)
    hojas_usadas = (db.session.query(func.sum(Venta.cantidad))
                    .filter_by(es_canje=False).scalar() or 0)
    restantes    = max(0, total_papel - hojas_usadas)
    minimo       = int(get_config('alerta_minimo', '100'))

    capacidad    = max(total_papel, minimo * 5) or 1
    porcentaje   = min(100, int(restantes / capacidad * 100))
    alerta       = restantes < minimo

    reposiciones = Inventario.query.order_by(Inventario.fecha.desc()).limit(30).all()

    return render_template('inventario.html',
                           restantes=restantes, hojas_usadas=hojas_usadas,
                           total_papel=total_papel, porcentaje=porcentaje,
                           alerta=alerta, minimo=minimo,
                           reposiciones=reposiciones)


@app.route('/inventario/agregar', methods=['POST'])
@admin_required
def agregar_inventario():
    tipo     = request.form.get('tipo', 'papel')
    cantidad = request.form.get('cantidad', type=int)
    nota     = request.form.get('nota', '').strip()

    if not cantidad or cantidad <= 0:
        flash('Cantidad inválida', 'error')
    else:
        db.session.add(Inventario(fecha=date.today(), tipo=tipo,
                                  cantidad=cantidad, nota=nota or None))
        db.session.commit()
        flash(f'Se registraron {cantidad} unidades de {tipo}', 'success')

    return redirect(url_for('inventario'))


# ============================================================
# SERVICIOS
# ============================================================

@app.route('/servicios')
@admin_required
def servicios():
    return render_template('servicios.html', servicios=Servicio.query.all())


@app.route('/servicios/editar/<int:id>', methods=['POST'])
@admin_required
def editar_servicio(id):
    s          = Servicio.query.get_or_404(id)
    precio     = request.form.get('precio',     type=float)
    costo_real = request.form.get('costo_real', type=float)
    if precio     is not None: s.precio     = precio
    if costo_real is not None: s.costo_real = costo_real
    db.session.commit()
    return jsonify({'success': True})


@app.route('/servicios/toggle/<int:id>', methods=['POST'])
@admin_required
def toggle_servicio(id):
    s        = Servicio.query.get_or_404(id)
    s.activo = not s.activo
    db.session.commit()
    return jsonify({'activo': s.activo})



# ============================================================
# GESTIÓN DE USUARIOS  (solo admin)
# ============================================================

@app.route('/usuarios')
@admin_required
def usuarios():
    lista = Usuario.query.order_by(Usuario.id).all()
    return render_template('usuarios.html', usuarios=lista)


@app.route('/usuarios/nuevo', methods=['POST'])
@admin_required
def nuevo_usuario():
    nombre = request.form.get('nombre', '').strip()
    pin    = request.form.get('pin', '').strip()
    rol    = request.form.get('rol', 'apoyo')

    if not nombre or not pin:
        return jsonify({'error': 'Nombre y PIN son requeridos'}), 400
    if len(pin) < 4 or not pin.isdigit():
        return jsonify({'error': 'El PIN debe ser numérico de al menos 4 dígitos'}), 400
    if Usuario.query.filter(db.func.lower(Usuario.nombre) == nombre.lower()).first():
        return jsonify({'error': f'Ya existe un usuario con el nombre "{nombre}"'}), 400
    if rol not in ('admin', 'apoyo'):
        rol = 'apoyo'

    u = Usuario(nombre=nombre, pin=generate_password_hash(pin), rol=rol)
    db.session.add(u)
    db.session.commit()
    return jsonify({'id': u.id, 'nombre': u.nombre, 'rol': u.rol})


@app.route('/usuarios/eliminar/<int:id>', methods=['POST'])
@admin_required
def eliminar_usuario(id):
    if id == session.get('usuario_id'):
        return jsonify({'error': 'No puedes eliminar tu propio usuario'}), 400
    u = Usuario.query.get_or_404(id)
    # Si tiene ventas, solo desactivar; si no, eliminar físicamente
    if u.ventas:
        u.activo = False
        db.session.commit()
        return jsonify({'accion': 'desactivado', 'nombre': u.nombre})
    else:
        db.session.delete(u)
        db.session.commit()
        return jsonify({'accion': 'eliminado', 'nombre': u.nombre})


@app.route('/usuarios/toggle/<int:id>', methods=['POST'])
@admin_required
def toggle_usuario(id):
    if id == session.get('usuario_id'):
        return jsonify({'error': 'No puedes desactivarte a ti mismo'}), 400
    u        = Usuario.query.get_or_404(id)
    u.activo = not u.activo
    db.session.commit()
    return jsonify({'activo': u.activo, 'nombre': u.nombre})

@app.route('/servicios/niveles/<int:id>')
@login_required
def get_niveles(id):
    niveles = ServicioNivel.query.filter_by(servicio_id=id).order_by(ServicioNivel.orden).all()
    srv     = Servicio.query.get_or_404(id)
    return jsonify({
        'niveles':          [{'id': n.id, 'nombre': n.nombre, 'precio': n.precio} for n in niveles],
        'precio_base':      srv.precio,
        'descuento_vol':    srv.descuento_volumen,
    })


@app.route('/servicios/niveles/agregar', methods=['POST'])
@admin_required
def agregar_nivel():
    servicio_id = request.form.get('servicio_id', type=int)
    nombre      = request.form.get('nombre', '').strip()
    precio      = request.form.get('precio', type=float)
    if not servicio_id or not nombre or precio is None:
        return jsonify({'error': 'Datos incompletos'}), 400
    orden = ServicioNivel.query.filter_by(servicio_id=servicio_id).count()
    n = ServicioNivel(servicio_id=servicio_id, nombre=nombre, precio=precio, orden=orden)
    db.session.add(n)
    db.session.commit()
    return jsonify({'id': n.id, 'nombre': n.nombre, 'precio': n.precio})


@app.route('/servicios/niveles/editar/<int:nivel_id>', methods=['POST'])
@admin_required
def editar_nivel(nivel_id):
    n      = ServicioNivel.query.get_or_404(nivel_id)
    nombre = request.form.get('nombre', '').strip()
    precio = request.form.get('precio', type=float)
    if nombre: n.nombre = nombre
    if precio is not None: n.precio = precio
    db.session.commit()
    return jsonify({'success': True})


@app.route('/servicios/niveles/eliminar/<int:nivel_id>', methods=['POST'])
@admin_required
def eliminar_nivel(nivel_id):
    n = ServicioNivel.query.get_or_404(nivel_id)
    db.session.delete(n)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/servicios/toggle_descuento/<int:id>', methods=['POST'])
@admin_required
def toggle_descuento(id):
    s = Servicio.query.get_or_404(id)
    s.descuento_volumen = not s.descuento_volumen
    db.session.commit()
    return jsonify({'descuento_volumen': s.descuento_volumen})


@app.route('/servicios/nuevo', methods=['POST'])
@admin_required
def nuevo_servicio():
    nombre     = request.form.get('nombre', '').strip()
    precio     = request.form.get('precio',     type=float)
    costo_real = request.form.get('costo_real', type=float)

    if not nombre:
        return jsonify({'error': 'El nombre es requerido'}), 400
    if precio is None or precio < 0:
        return jsonify({'error': 'El precio no es válido'}), 400
    # precio = 0 significa precio variable (se ingresa en cada venta)
    if costo_real is None or costo_real < 0:
        costo_real = 0.0  # opcional: sin costo = margen desconocido

    desc_vol = request.form.get('descuento_volumen', 'false') == 'true'
    s = Servicio(nombre=nombre, precio=precio, costo_real=costo_real,
                 activo=True, descuento_volumen=desc_vol)
    db.session.add(s)
    db.session.flush()
    # Niveles iniciales enviados como JSON
    import json as _json
    niveles_json = request.form.get('niveles', '[]')
    try:
        niveles_data = _json.loads(niveles_json)
        for i, niv in enumerate(niveles_data):
            if niv.get('nombre') and niv.get('precio') is not None:
                db.session.add(ServicioNivel(
                    servicio_id=s.id, nombre=niv['nombre'],
                    precio=float(niv['precio']), orden=i))
    except Exception:
        pass
    db.session.commit()
    margen = ((precio - costo_real) / precio * 100) if precio > 0 else 0
    return jsonify({'id': s.id, 'nombre': s.nombre, 'precio': s.precio,
                    'costo_real': s.costo_real, 'margen': round(margen, 1),
                    'descuento_volumen': s.descuento_volumen})


@app.route('/servicios/eliminar/<int:id>', methods=['POST'])
@admin_required
def eliminar_servicio(id):
    s = Servicio.query.get_or_404(id)
    if s.ventas:
        # Tiene historial: solo desactivar para preservar datos
        s.activo = False
        db.session.commit()
        return jsonify({'accion': 'desactivado', 'nombre': s.nombre})
    else:
        db.session.delete(s)
        db.session.commit()
        return jsonify({'accion': 'eliminado', 'nombre': s.nombre})




# ============================================================
# EXPORTAR EXCEL
# ============================================================

@app.route('/exportar')
@admin_required
def exportar():
    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')

    q = Venta.query
    if desde:
        q = q.filter(Venta.fecha >= datetime.strptime(desde, '%Y-%m-%d').date())
    if hasta:
        q = q.filter(Venta.fecha <= datetime.strptime(hasta, '%Y-%m-%d').date())
    ventas = q.order_by(Venta.fecha, Venta.hora).all()

    nombre_negocio  = get_config('nombre_negocio', 'Mi Copiadora')
    costos_fijos    = float(get_config('costos_fijos_mes', '160'))

    AZUL   = 'FF185FA5'
    VERDE  = 'FF1D9E75'
    GRIS   = 'FFF2F2F2'
    BLANCO = 'FFFFFFFF'

    total_ingresos = sum(v.total for v in ventas)
    total_costos   = sum(v.servicio.costo_real * v.cantidad for v in ventas)
    util_bruta     = total_ingresos - total_costos
    total_tx       = len(ventas)
    ticket_prom    = total_ingresos / total_tx if total_tx else 0
    margen_bruto   = (util_bruta / total_ingresos * 100) if total_ingresos else 0

    wb  = Workbook()

    # ── HOJA 1: RESUMEN ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Resumen'
    periodo   = f"{desde or 'Inicio'} al {hasta or str(date.today())}"

    filas_res = [
        ('RESUMEN DEL PERÍODO', None, True, AZUL),
        ('Negocio:', nombre_negocio, False, None),
        ('Período:', periodo, False, None),
        (None, None, False, None),
        ('MÉTRICAS PRINCIPALES', None, True, AZUL),
        ('Total de ingresos:', total_ingresos, False, None),
        ('Total costos estimados:', total_costos, False, None),
        ('Utilidad bruta:', util_bruta, False, None),
        ('Margen bruto promedio:', f'{margen_bruto:.1f}%', False, None),
        ('Transacciones totales:', total_tx, False, None),
        ('Ticket promedio:', ticket_prom, False, None),
    ]
    for i, (a, b, header, color) in enumerate(filas_res, 1):
        c1 = ws1.cell(i, 1, a)
        c2 = ws1.cell(i, 2, b)
        if header:
            c1.font  = Font(bold=True, color=BLANCO, size=11)
            c1.fill  = PatternFill('solid', fgColor=color)
            ws1.merge_cells(f'A{i}:B{i}')
        elif a and a.startswith('Total') or (a and 'Ticket' in a) or (a and 'Utilidad' in a):
            c1.font = Font(bold=True)
            if isinstance(b, (int, float)):
                c2.number_format = '"S/"#,##0.00'
    ws1.column_dimensions['A'].width = 32
    ws1.column_dimensions['B'].width = 22

    # ── HOJA 2: VENTAS POR DÍA ───────────────────────────────
    ws2 = wb.create_sheet('Ventas por día')
    hdrs2 = ['Fecha', 'Transacciones', 'Ingresos (S/)', 'Costos Est. (S/)', 'Utilidad (S/)']
    for j, h in enumerate(hdrs2, 1):
        c = ws2.cell(1, j, h)
        c.font = Font(bold=True, color=BLANCO)
        c.fill = PatternFill('solid', fgColor=VERDE)

    por_dia = defaultdict(list)
    for v in ventas:
        por_dia[v.fecha].append(v)

    for i, dia in enumerate(sorted(por_dia), 2):
        lst = por_dia[dia]
        ing = sum(x.total for x in lst)
        cos = sum(x.servicio.costo_real * x.cantidad for x in lst)
        row = [dia.strftime('%d/%m/%Y'), len(lst), ing, cos, ing - cos]
        fill = PatternFill('solid', fgColor=GRIS) if i % 2 == 0 else None
        for j, val in enumerate(row, 1):
            c = ws2.cell(i, j, val)
            if fill: c.fill = fill
            if j in (3, 4, 5): c.number_format = '"S/"#,##0.00'
    for j, w in enumerate([15, 15, 16, 16, 16], 1):
        ws2.column_dimensions[get_column_letter(j)].width = w

    # ── HOJA 3: VENTAS POR SERVICIO ──────────────────────────
    ws3 = wb.create_sheet('Ventas por servicio')
    hdrs3 = ['Servicio', 'Unidades', 'Ingresos (S/)', 'Costo Total (S/)', 'Utilidad (S/)', 'Margen %']
    for j, h in enumerate(hdrs3, 1):
        c = ws3.cell(1, j, h)
        c.font = Font(bold=True, color=BLANCO)
        c.fill = PatternFill('solid', fgColor=AZUL)

    por_srv = defaultdict(list)
    for v in ventas:
        por_srv[v.servicio.nombre].append(v)

    for i, srv_nombre in enumerate(sorted(por_srv), 2):
        lst = por_srv[srv_nombre]
        uni = sum(x.cantidad for x in lst)
        ing = sum(x.total for x in lst)
        cos = sum(x.servicio.costo_real * x.cantidad for x in lst)
        uti = ing - cos
        mgn = (uti / ing * 100) if ing else 0
        row = [srv_nombre, uni, ing, cos, uti, f'{mgn:.1f}%']
        fill = PatternFill('solid', fgColor=GRIS) if i % 2 == 0 else None
        for j, val in enumerate(row, 1):
            c = ws3.cell(i, j, val)
            if fill: c.fill = fill
            if j in (3, 4, 5): c.number_format = '"S/"#,##0.00'
    for j, w in enumerate([26, 10, 15, 16, 15, 12], 1):
        ws3.column_dimensions[get_column_letter(j)].width = w

    # ── HOJA 4: DETALLE COMPLETO ─────────────────────────────
    ws4 = wb.create_sheet('Detalle completo')
    hdrs4 = ['Fecha', 'Hora', 'Servicio', 'Cantidad', 'Precio Unit.', 'Total', 'Usuario', 'Cliente', 'Canje']
    for j, h in enumerate(hdrs4, 1):
        c = ws4.cell(1, j, h)
        c.font = Font(bold=True, color=BLANCO)
        c.fill = PatternFill('solid', fgColor=VERDE)

    for i, v in enumerate(ventas, 2):
        row = [
            v.fecha.strftime('%d/%m/%Y'),
            v.hora.strftime('%H:%M'),
            v.servicio.nombre, v.cantidad,
            v.precio_unitario, v.total,
            v.usuario.nombre,
            v.cliente.nombre if v.cliente else '—',
            'Sí' if v.es_canje else 'No',
        ]
        fill = PatternFill('solid', fgColor=GRIS) if i % 2 == 0 else None
        for j, val in enumerate(row, 1):
            c = ws4.cell(i, j, val)
            if fill: c.fill = fill
            if j in (5, 6): c.number_format = '"S/"#,##0.00'
    for j, w in enumerate([14, 8, 24, 10, 14, 14, 12, 20, 8], 1):
        ws4.column_dimensions[get_column_letter(j)].width = w

    # ── HOJA 5: RENTABILIDAD ─────────────────────────────────
    ws5 = wb.create_sheet('Rentabilidad')
    util_neta  = util_bruta - costos_fijos
    margen_neto= (util_neta / total_ingresos * 100) if total_ingresos else 0

    if margen_neto > 60:
        salud = 'SALUDABLE ✓'
        color_salud = VERDE
        recom = 'Excelente estado. Considera expandir servicios o adquirir mejor equipo.'
    elif margen_neto >= 30:
        salud = 'ATENCIÓN ⚠'
        color_salud = 'FFFFC000'
        recom = 'Negocio rentable con margen de mejora. Revisa costos y busca aumentar ventas.'
    else:
        salud = 'REVISAR ✗'
        color_salud = 'FFFF4444'
        recom = 'Atención urgente. Analiza tus gastos y considera ajustar precios o reducir costos.'

    filas5 = [
        ('ANÁLISIS DE RENTABILIDAD', None, AZUL),
        (None, None, None),
        ('RESULTADOS FINANCIEROS', None, AZUL),
        ('Ingresos totales del período:', total_ingresos, None),
        ('Costos variables estimados:', total_costos, None),
        ('Costos fijos del mes:', costos_fijos, None),
        ('Utilidad neta estimada:', util_neta, None),
        ('Margen neto:', f'{margen_neto:.1f}%', None),
        (None, None, None),
        ('DISTRIBUCIÓN RECOMENDADA DE UTILIDAD', None, VERDE),
        ('40% — Reinversión en insumos (papel, tinta):', util_neta * 0.40, None),
        ('30% — Ahorro / Fondo de emergencia:', util_neta * 0.30, None),
        ('20% — Crecimiento (nuevos servicios, equipo):', util_neta * 0.20, None),
        ('10% — Personal / Retiro propio:', util_neta * 0.10, None),
        (None, None, None),
        ('INDICADOR DE SALUD DEL NEGOCIO', salud, color_salud),
        ('Recomendación:', recom, None),
    ]
    for i, (a, b, color) in enumerate(filas5, 1):
        c1 = ws5.cell(i, 1, a)
        c2 = ws5.cell(i, 2, b)
        if color:
            c1.font = Font(bold=True, color=BLANCO, size=11)
            c1.fill = PatternFill('solid', fgColor=color)
            if i == 16:
                c2.font = Font(bold=True, color=BLANCO)
                c2.fill = PatternFill('solid', fgColor=color)
            else:
                ws5.merge_cells(f'A{i}:B{i}')
        if isinstance(b, float):
            c2.number_format = '"S/"#,##0.00'

    ws5.column_dimensions['A'].width = 44
    ws5.column_dimensions['B'].width = 24

    # Enviar archivo
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"reporte_{nombre_negocio.replace(' ', '_')}_{date.today()}.xlsx"
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# EDITAR / ELIMINAR VENTAS  (solo admin)
# ============================================================

@app.route('/ventas/editar/<int:id>', methods=['POST'])
@admin_required
def editar_venta(id):
    v           = Venta.query.get_or_404(id)
    cantidad    = request.form.get('cantidad',    type=int)
    precio      = request.form.get('precio',      type=float)
    descripcion = request.form.get('descripcion', '').strip()
    cliente_id  = request.form.get('cliente_id', type=int)
    quitar_cli  = request.form.get('quitar_cliente', 'false') == 'true'

    if cantidad and cantidad > 0:
        v.cantidad = cantidad
    if precio is not None and precio >= 0:
        v.precio_unitario = precio
    if descripcion:
        v.descripcion = descripcion

    # Manejo de cliente
    if quitar_cli:
        v.cliente_id = None
    elif cliente_id:
        v.cliente_id = cliente_id

    v.total = round(v.precio_unitario * v.cantidad, 2)

    # Registrar puntos si hay cliente
    if v.cliente_id and not v.es_canje:
        cli = Cliente.query.get(v.cliente_id)
        if cli:
            puntos = floor(v.total / 10)
            if puntos > 0:
                cli.puntos += puntos
                db.session.add(PuntosHistorial(
                    cliente_id=v.cliente_id, puntos=puntos,
                    descripcion=f'Corrección de venta #{id}'))

    db.session.commit()
    return jsonify({
        'success':         True,
        'total':           v.total,
        'cantidad':        v.cantidad,
        'precio_unitario': v.precio_unitario,
        'cliente_id':      v.cliente_id,
    })


@app.route('/ventas/eliminar/<int:id>', methods=['POST'])
@admin_required
def eliminar_venta(id):
    v = Venta.query.get_or_404(id)
    # Revertir puntos si tiene cliente
    if v.cliente_id and not v.es_canje:
        cli = Cliente.query.get(v.cliente_id)
        if cli:
            puntos_restar = floor(v.total / 10)
            cli.puntos = max(0, cli.puntos - puntos_restar)
            if puntos_restar > 0:
                db.session.add(PuntosHistorial(
                    cliente_id=v.cliente_id, puntos=-puntos_restar,
                    descripcion=f'Venta #{id} eliminada'))
    db.session.delete(v)
    db.session.commit()
    return jsonify({'success': True})


# ============================================================
# CRM
# ============================================================

@app.route('/crm')
@admin_required
def crm():
    from sqlalchemy import func
    clientes = Cliente.query.filter_by(activo=True).order_by(Cliente.nombre).all()

    frecuentes = (db.session.query(
                      Cliente,
                      func.count(Venta.id).label('num_ventas'),
                      func.sum(Venta.total).label('total_gastado'))
                  .join(Venta, Venta.cliente_id == Cliente.id)
                  .group_by(Cliente.id)
                  .order_by(func.count(Venta.id).desc())
                  .limit(10).all())

    return render_template('crm.html', clientes=clientes, frecuentes=frecuentes)


@app.route('/crm/cliente/nuevo', methods=['POST'])
@admin_required
def nuevo_cliente():
    nombre = request.form.get('nombre', '').strip()
    celular= request.form.get('celular', '').strip()
    if not nombre:
        return jsonify({'error': 'El nombre es requerido'}), 400
    c = Cliente(nombre=nombre, celular=celular or None)
    db.session.add(c)
    db.session.commit()
    return jsonify({'id': c.id, 'nombre': c.nombre,
                    'celular': c.celular or '', 'puntos': 0})


@app.route('/crm/cliente/<int:id>')
@admin_required
def ver_cliente(id):
    cliente  = Cliente.query.get_or_404(id)
    historial= (Venta.query.filter_by(cliente_id=id)
                .order_by(Venta.fecha.desc(), Venta.hora.desc()).all())
    return render_template('crm_cliente.html', cliente=cliente, historial=historial)


@app.route('/crm/cliente/<int:id>/editar', methods=['POST'])
@admin_required
def editar_cliente(id):
    c        = Cliente.query.get_or_404(id)
    c.nombre = request.form.get('nombre', c.nombre).strip()
    c.celular= request.form.get('celular', '').strip() or None
    db.session.commit()
    return jsonify({'success': True})


@app.route('/crm/buscar')
@login_required
def buscar_cliente():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    resultados = Cliente.query.filter(
        Cliente.activo == True,
        db.or_(Cliente.nombre.ilike(f'%{q}%'), Cliente.celular.ilike(f'%{q}%'))
    ).limit(10).all()
    return jsonify([{'id': c.id, 'nombre': c.nombre,
                     'celular': c.celular or '', 'puntos': c.puntos}
                    for c in resultados])


# ============================================================
# INICIO
# ============================================================

with app.app_context():
    init_db()

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0',
            port=int(os.environ.get('PORT', 5000)))
